import os
import logging
import csv
import hashlib
import time
from openpyxl import load_workbook
from celery import shared_task
from celery.exceptions import SoftTimeLimitExceeded, MaxRetriesExceededError, Retry
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.core.cache import cache

from .service_common import upload_batch_key

logger = logging.getLogger(__name__)


def _send_ws(user_id, message, status):
    """Send a WebSocket progress message to the user's channel group."""
    try:
        # Prevent flooding Redis channel queues with very frequent progress pings.
        # Final states should always be delivered immediately.
        if status == "processing":
            msg_hash = hashlib.md5(str(message).encode("utf-8")).hexdigest()
            dedupe_key = f"upload_ws_dedupe_{user_id}_{status}_{msg_hash}"
            # Skip repeated identical processing message for a short window.
            if not cache.add(dedupe_key, 1, timeout=2):
                return

            # Global per-user processing throttle (~2 messages / second).
            # Use a time-bucket key to avoid sub-second cache timeout precision issues.
            throttle_bucket = int(time.monotonic() * 2)
            throttle_key = f"upload_ws_throttle_{user_id}_{throttle_bucket}"
            if not cache.add(throttle_key, 1, timeout=2):
                return

        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            f"user_{user_id}",
            {
                "type": "upload_progress",
                "message": message,
                "status": status,
            },
        )
    except Exception as exc:
        logger.warning("[UploadTask] WebSocket send failed: %s", exc)


def _enqueue_dashboard_warmup(
    data_owner_id,
    *,
    skip_during_upload=True,
    filter_sets=None,
    view_types=None,
):
    try:
        from django.conf import settings

        if not getattr(settings, "DASHBOARD_WARMUP_ENABLED", True):
            return
        # During uploads the warmup is skipped by default — it computes up to
        # 21 expensive dashboard payloads that are rarely needed immediately.
        # The first dashboard page-load will compute and cache on demand.
        if skip_during_upload:
            logger.info(
                "[UploadTask] Skipping warmup for user=%s (skip_during_upload=True)",
                data_owner_id,
            )
            return
        from apps.dashboard.tasks import warmup_dashboard_payloads_task

        warmup_dashboard_payloads_task.delay(
            data_owner_id=data_owner_id,
            filter_sets=filter_sets,
            view_types=view_types,
        )
    except Exception:
        logger.exception(
            "[UploadTask] Failed to enqueue dashboard warmup for user=%s", data_owner_id
        )


def _enqueue_daily_summary_refresh(data_owner_id, affected_dates=None):
    try:
        from django.conf import settings

        if not getattr(settings, "DASHBOARD_DAILY_SUMMARY_ENABLED", True):
            return
        from apps.dashboard.tasks import refresh_dashboard_daily_summary_task

        refresh_dashboard_daily_summary_task.delay(
            data_owner_id=data_owner_id,
            only_dates=list(affected_dates or []),
        )
    except Exception:
        logger.exception(
            "[UploadTask] Failed to enqueue daily summary refresh for user=%s",
            data_owner_id,
        )


def _enqueue_inventory_summary_refresh(data_owner_id, affected_dates=None):
    try:
        from django.conf import settings

        if not getattr(settings, "DASHBOARD_INVENTORY_SUMMARY_ENABLED", True):
            return
        from apps.dashboard.tasks import refresh_dashboard_inventory_summary_task

        refresh_dashboard_inventory_summary_task.delay(
            data_owner_id=data_owner_id,
            only_dates=list(affected_dates or []),
        )
    except Exception:
        logger.exception(
            "[UploadTask] Failed to enqueue inventory summary refresh for user=%s",
            data_owner_id,
        )


def _enqueue_asin_monthly_summary_refresh(data_owner_id, affected_dates=None):
    try:
        from django.conf import settings

        if not getattr(settings, "DASHBOARD_ASIN_MONTHLY_SUMMARY_ENABLED", True):
            return
        from apps.dashboard.tasks import refresh_dashboard_asin_monthly_summary_task

        affected_months = sorted({
            str(d)[:7] + "-01" for d in (affected_dates or [])
        })
        refresh_dashboard_asin_monthly_summary_task.delay(
            data_owner_id=data_owner_id,
            only_months=affected_months,
        )
    except Exception:
        logger.exception(
            "[UploadTask] Failed to enqueue ASIN monthly summary refresh for user=%s",
            data_owner_id,
        )


def _dashboard_refresh_cache_key(data_owner_id):
    return f"dashboard_refresh_status_{data_owner_id}"


def _set_dashboard_refresh_status(data_owner_id, state, message, timeout=3600):
    cache.set(
        _dashboard_refresh_cache_key(data_owner_id),
        {
            "state": state,
            "message": message,
            "updated_at_ts": time.time(),
        },
        timeout=timeout,
    )


def _mark_batch_task_complete(
    *,
    batch_id,
    batch_total,
    user_id,
    data_owner_id,
    is_flipkart,
    success,
    file_type="",
    affected_dates=None,
):
    if not batch_id:
        return

    ttl = 86400
    expected_key = upload_batch_key(batch_id, "expected_total")
    completed_key = upload_batch_key(batch_id, "completed_total")
    failed_key = upload_batch_key(batch_id, "failed_total")
    finalized_key = upload_batch_key(batch_id, "finalized")
    meta_key = upload_batch_key(batch_id, "meta")
    file_types_key = upload_batch_key(batch_id, "file_types")
    dates_key = upload_batch_key(batch_id, "affected_dates")

    # Accumulate file types and affected dates from each task.
    # Note: get-modify-set has a small race window with concurrent workers.
    # Worst case: a file_type or date is lost, causing a full rebuild instead
    # of incremental — safe but slower. We retry once to minimise this.
    if file_type:
        for _attempt in range(3):
            existing_types = cache.get(file_types_key) or set()
            existing_types.add(file_type)
            cache.set(file_types_key, existing_types, timeout=ttl)
            # Verify the write stuck
            verify = cache.get(file_types_key)
            if verify and file_type in verify:
                break

    if affected_dates:
        for _attempt in range(3):
            existing_dates = cache.get(dates_key) or set()
            for d in affected_dates:
                existing_dates.add(str(d))
            cache.set(dates_key, existing_dates, timeout=ttl)
            break  # dates are additive; minor race loss is acceptable

    expected = cache.get(expected_key)
    if expected is None:
        expected = int(batch_total or 0)
        if expected > 0:
            cache.set(expected_key, expected, timeout=ttl)
            logger.info(
                "[BatchTracker] batch=%s expected_total key was missing, re-set to %d",
                batch_id, expected,
            )
    else:
        expected = int(expected)
    if expected <= 0:
        logger.warning(
            "[BatchTracker] batch=%s expected_total=0, skipping finalization check. "
            "batch_total kwarg=%s",
            batch_id, batch_total,
        )
        return

    # Atomically increment the completed counter.
    # Use add-then-incr to avoid the race where two workers both fallback
    # to cache.set(key, 1) and one increment is lost.
    try:
        completed = cache.incr(completed_key)
    except ValueError:
        # Key doesn't exist yet — seed it at 0, then incr.
        cache.add(completed_key, 0, timeout=ttl)
        try:
            completed = cache.incr(completed_key)
        except ValueError:
            # Extremely unlikely double-race fallback.
            cache.set(completed_key, 1, timeout=ttl)
            completed = 1

    logger.info(
        "[BatchTracker] batch=%s completed=%d/%d success=%s",
        batch_id, completed, expected, success,
    )

    if not success:
        try:
            cache.incr(failed_key)
        except ValueError:
            cache.add(failed_key, 0, timeout=ttl)
            try:
                cache.incr(failed_key)
            except ValueError:
                cache.set(failed_key, 1, timeout=ttl)

    if completed < expected:
        return

    # Finalize only once, even if multiple workers reach completion concurrently.
    if not cache.add(finalized_key, 1, timeout=ttl):
        logger.info(
            "[BatchTracker] batch=%s already finalized by another worker, skipping.",
            batch_id,
        )
        return

    failed_total = int(cache.get(failed_key) or 0)
    meta = cache.get(meta_key) or {}
    owner_id = int(meta.get("data_owner_id") or data_owner_id)
    owner_user_id = int(meta.get("user_id") or user_id)
    owner_is_flipkart = bool(meta.get("is_flipkart")) if "is_flipkart" in meta else bool(is_flipkart)

    # Determine smart file_type and affected_dates from accumulated batch data.
    batch_file_types = cache.get(file_types_key) or set()
    batch_affected_dates = sorted(cache.get(dates_key) or set())

    logger.info(
        "[BatchTracker] batch=%s FINALIZING — completed=%d expected=%d failed=%d owner=%d flipkart=%s",
        batch_id, completed, expected, failed_total, owner_id, owner_is_flipkart,
    )

    if failed_total > 0:
        msg = (
            "We could not process a few uploaded files, so the dashboard was not refreshed. "
            "Please re-upload the failed files and try again."
        )
        _set_dashboard_refresh_status(owner_id, "error", msg, timeout=900)
        _send_ws(owner_user_id, msg, "error")
        return

    # Determine the effective file_type for the refresh.
    # Only force full rebuild when category/price files are in the batch.
    # For sales-only or spend-only batches, enable incremental date-scoped refresh.
    FULL_REBUILD_TYPES_AMZ = {"category", "price"}
    FULL_REBUILD_TYPES_FK = {"fk_category", "fk_price"}
    full_rebuild_types = FULL_REBUILD_TYPES_FK if owner_is_flipkart else FULL_REBUILD_TYPES_AMZ

    needs_full_rebuild = bool(batch_file_types & full_rebuild_types)
    if needs_full_rebuild:
        effective_file_type = "fk_category" if owner_is_flipkart else "category"
        effective_dates = []  # full rebuild ignores dates
    else:
        # Pick the dominant file type; prefer "sales"/"spend" for incremental.
        if owner_is_flipkart:
            if "fk_search_traffic" in batch_file_types:
                effective_file_type = "fk_search_traffic"
            elif "fk_pla" in batch_file_types:
                effective_file_type = "fk_pla"
            else:
                effective_file_type = next(iter(batch_file_types), "fk_category")
        else:
            if "sales" in batch_file_types:
                effective_file_type = "sales"
            elif "spend" in batch_file_types:
                effective_file_type = "spend"
            else:
                effective_file_type = next(iter(batch_file_types), "category")
        effective_dates = batch_affected_dates

    _set_dashboard_refresh_status(
        owner_id,
        "processing",
        "Dashboard updating in process as per the new uploaded data.",
        timeout=3600,
    )
    _send_ws(
        owner_user_id,
        "All files uploaded successfully. Dashboard update started.",
        "partial",
    )
    refresh_dashboard_after_upload_task.delay(
        data_owner_id=owner_id,
        user_id=owner_user_id,
        is_flipkart=owner_is_flipkart,
        file_type=effective_file_type,
        affected_dates=effective_dates,
        dashboard_refreshed=False,
    )


def _run_dashboard_refresh(
    *,
    data_owner,
    user_id,
    is_flipkart,
    file_type,
    affected_dates=None,
    dashboard_refreshed=False,
):
    from apps.upload.dashboard_builders import (
        generate_dashboard_data,
        generate_flipkart_dashboard_data,
        update_category_in_processed_data,
        update_price_in_processed_data,
        update_fk_category_in_processed_data,
        update_fk_price_in_processed_data,
    )
    from apps.dashboard.services.invalidation import invalidate_dashboard_cache_for_user
    from apps.dashboard.models import (
        Flipkartfba,
        FlipkartCategoryMap,
        FlipkartSearchTraffic,
        FlipkartPLA,
        FlipkartPrice,
        FlipkartInventoryStock,
    )

    affected_dates = set(affected_dates or [])
    dashboard_invalidated = False
    _t_start = time.monotonic()

    def _dashboard_progress(message):
        _send_ws(user_id, message, "processing")

    if is_flipkart:
        # Batch all 6 existence checks into one query per table using
        # a single COUNT per model — avoids 6 sequential DB round trips.
        uid = data_owner.id
        has_fk_category = FlipkartCategoryMap.objects.filter(user_id=uid).values("id")[:1].exists()
        has_fk_traffic = FlipkartSearchTraffic.objects.filter(user_id=uid).values("id")[:1].exists()
        has_fk_pla = FlipkartPLA.objects.filter(user_id=uid).values("id")[:1].exists()
        has_fk_price = FlipkartPrice.objects.filter(user_id=uid).values("id")[:1].exists()
        if not (has_fk_category and has_fk_traffic and has_fk_pla and has_fk_price):
            raise ValueError(
                "Flipkart requires Search Traffic, Category, PLA, and Price reports."
            )

        has_fba_stock = Flipkartfba.objects.filter(user_id=uid).values("id")[:1].exists()
        has_fk_inventory = FlipkartInventoryStock.objects.filter(user_id=uid).values("id")[:1].exists()
        if not has_fba_stock or not has_fk_inventory:
            logger.warning(
                "[DashboardRefresh] user=%s missing FK Inventory (%s) or FBA Stock (%s) — "
                "skipping inventory health, but will still build core dashboard.",
                data_owner.id, has_fk_inventory, has_fba_stock,
            )

        if file_type in {"fk_search_traffic", "fk_pla"} and affected_dates:
            generate_flipkart_dashboard_data(
                data_owner,
                progress_callback=_dashboard_progress,
                only_dates=sorted(affected_dates),
            )
            dashboard_refreshed = True
        elif file_type == "fk_category":
            update_fk_category_in_processed_data(data_owner.id)
            invalidate_dashboard_cache_for_user(data_owner.id, clear_materialized=True)
            dashboard_refreshed = True
        elif file_type == "fk_price":
            update_fk_price_in_processed_data(data_owner.id)
            invalidate_dashboard_cache_for_user(data_owner.id, clear_materialized=True)
            dashboard_refreshed = True
        else:
            invalidate_dashboard_cache_for_user(data_owner.id, clear_materialized=True)
            dashboard_invalidated = True
    elif file_type in {"sales", "spend"}:
        if affected_dates:
            generate_dashboard_data(
                data_owner,
                progress_callback=_dashboard_progress,
                only_dates=sorted(affected_dates),
            )
        else:
            generate_dashboard_data(data_owner, progress_callback=_dashboard_progress)
        dashboard_refreshed = True
    elif file_type == "category":
        update_category_in_processed_data(data_owner.id)
        invalidate_dashboard_cache_for_user(data_owner.id, clear_materialized=True)
        dashboard_refreshed = True
    elif file_type == "price":
        update_price_in_processed_data(data_owner.id)
        invalidate_dashboard_cache_for_user(data_owner.id, clear_materialized=True)
        dashboard_refreshed = True
    else:
        invalidate_dashboard_cache_for_user(data_owner.id, clear_materialized=True)
        dashboard_invalidated = True

    _t_dashboard = time.monotonic()
    logger.info(
        "[DashboardRefreshTask] Dashboard rebuild completed in %.1fs for owner=%s",
        _t_dashboard - _t_start, data_owner.id,
    )

    if dashboard_refreshed or dashboard_invalidated:
        sorted_dates = sorted(affected_dates or [])
        affected_months = sorted({
            str(d)[:7] + "-01" for d in sorted_dates
        }) if sorted_dates else []

        # Rebuild DashboardDailySummary INLINE (synchronously) so the fast
        # summary rollup path is ready before the user is notified.
        # This is the key table that makes filter changes fast — without it
        # every dashboard load falls back to the 17s GROUP BY asin query.
        try:
            from apps.dashboard.services.daily_summary import rebuild_daily_summary_for_user
            _t_ds = time.monotonic()
            rebuild_daily_summary_for_user(data_owner, only_dates=sorted_dates or None)
            logger.info(
                "[DashboardRefreshTask] Daily summary rebuilt inline in %.1fs for owner=%s",
                time.monotonic() - _t_ds, data_owner.id,
            )
        except Exception:
            logger.exception(
                "[DashboardRefreshTask] Inline daily summary rebuild failed for owner=%s, "
                "falling back to async task.", data_owner.id,
            )
            _enqueue_daily_summary_refresh(data_owner.id, affected_dates=sorted_dates)

        # Fire inventory + monthly summaries as async tasks — they're not needed
        # for the first dashboard open and can complete in the background.
        try:
            from celery import group as celery_group
            from apps.dashboard.tasks import (
                refresh_dashboard_inventory_summary_task,
                refresh_dashboard_asin_monthly_summary_task,
            )
            bg_group = celery_group(
                refresh_dashboard_inventory_summary_task.si(
                    data_owner_id=data_owner.id,
                    only_dates=list(sorted_dates),
                ),
                refresh_dashboard_asin_monthly_summary_task.si(
                    data_owner_id=data_owner.id,
                    only_months=affected_months,
                ),
            )
            bg_group.apply_async()
            logger.info(
                "[DashboardRefreshTask] Dispatched inventory+asin-monthly summary tasks "
                "in background for owner=%s (%d dates, %d months)",
                data_owner.id, len(sorted_dates), len(affected_months),
            )
        except Exception:
            logger.exception(
                "[DashboardRefreshTask] Background summary dispatch failed for owner=%s",
                data_owner.id,
            )
            _enqueue_inventory_summary_refresh(data_owner.id, affected_dates=sorted_dates)
            _enqueue_asin_monthly_summary_refresh(data_owner.id, affected_dates=sorted_dates)

        # Prime the default dashboard payload right after refresh so the first
        # post-upload page open can reuse cache instead of recomputing from scratch.
        _enqueue_dashboard_warmup(
            data_owner.id,
            skip_during_upload=False,
            filter_sets=[{}],
        )


@shared_task(bind=True)
def refresh_dashboard_after_upload_task(
    self,
    data_owner_id,
    user_id,
    is_flipkart,
    file_type,
    affected_dates=None,
    dashboard_refreshed=False,
):
    from apps.accounts.models import Users

    try:
        lock_key = f"dashboard_refresh_lock_{data_owner_id}"
        lock_token = self.request.id

        if not cache.add(lock_key, lock_token, timeout=1800):
            # Check if the existing lock is stale (older than 15 minutes).
            existing_token = cache.get(lock_key)
            lock_status = cache.get(f"{lock_key}_ts")
            stale_threshold = 900  # 15 minutes
            if lock_status and (time.time() - float(lock_status)) > stale_threshold:
                logger.warning(
                    "[DashboardRefreshTask] Stale lock detected for owner=%s "
                    "(held by %s for %.0fs). Forcing release.",
                    data_owner_id, existing_token,
                    time.time() - float(lock_status),
                )
                cache.delete(lock_key)
                cache.delete(f"{lock_key}_ts")
                # Try to acquire again after clearing stale lock.
                if not cache.add(lock_key, lock_token, timeout=1800):
                    logger.info(
                        "[DashboardRefreshTask] Lock re-acquired by another worker after stale clear. Retrying."
                    )
                    raise self.retry(countdown=10, max_retries=30)
            else:
                logger.info(
                    "[DashboardRefreshTask] Lock held by %s for owner=%s. "
                    "Retry %d. task=%s",
                    existing_token, data_owner_id,
                    self.request.retries, lock_token,
                )
                raise self.retry(countdown=10, max_retries=30)

        # Record lock acquisition time for staleness detection.
        cache.set(f"{lock_key}_ts", str(time.time()), timeout=1800)

        data_owner = Users.objects.get(pk=data_owner_id)
        _refresh_t0 = time.time()
        _dates_count = len(affected_dates or [])
        logger.info(
            "[DashboardRefreshTask] Starting refresh for owner=%s flipkart=%s file_type=%s dates=%d",
            data_owner_id, is_flipkart, file_type, _dates_count,
        )
        _set_dashboard_refresh_status(
            data_owner_id,
            "processing",
            "Dashboard updating in process as per the new uploaded data.",
            timeout=3600,
        )
        _send_ws(
            user_id,
            "Dashboard updating in process as per the new uploaded data.",
            "processing",
        )

        _run_dashboard_refresh(
            data_owner=data_owner,
            user_id=user_id,
            is_flipkart=is_flipkart,
            file_type=file_type,
            affected_dates=affected_dates or [],
            dashboard_refreshed=dashboard_refreshed,
        )

        _refresh_elapsed = time.time() - _refresh_t0
        _set_dashboard_refresh_status(
            data_owner_id,
            "success",
            "Dashboard updated with the new updated data.",
            timeout=300,
        )
        _send_ws(user_id, "Dashboard updated with the new updated data.", "complete")
        logger.info(
            "[DashboardRefreshTask] Completed successfully for owner=%s in %.1fs (dates=%d)",
            data_owner_id, _refresh_elapsed, _dates_count,
        )
        return {"status": "success"}
    except Retry:
        raise
    except MaxRetriesExceededError:
        logger.error(
            "[DashboardRefreshTask] Max retries exceeded for owner=%s. "
            "Lock may be permanently stuck. Clearing lock and notifying user.",
            data_owner_id,
        )
        # Force-clear the lock so subsequent uploads aren't permanently blocked.
        try:
            cache.delete(f"dashboard_refresh_lock_{data_owner_id}")
            cache.delete(f"dashboard_refresh_lock_{data_owner_id}_ts")
        except Exception:
            pass
        msg = "Dashboard update timed out waiting for a previous update to finish. Please try uploading again."
        _set_dashboard_refresh_status(data_owner_id, "error", msg, timeout=600)
        _send_ws(user_id, msg, "error")
        return {"status": "error", "message": msg}
    except Exception as exc:
        logger.exception("[DashboardRefreshTask] Error for owner=%s: %s", data_owner_id, exc)
        _set_dashboard_refresh_status(
            data_owner_id,
            "error",
            f"Dashboard update failed: {str(exc)}",
            timeout=600,
        )
        _send_ws(user_id, f"Dashboard update failed: {str(exc)}", "error")
        return {"status": "error", "message": str(exc)}
    finally:
        try:
            lock_key = f"dashboard_refresh_lock_{data_owner_id}"
            if cache.get(lock_key) == self.request.id:
                cache.delete(lock_key)
                cache.delete(f"{lock_key}_ts")
        except Exception:
            pass


NON_CONVERTIBLE_EXCEL_TYPES = set()
EXCEL_TO_CSV_MIN_SIZE_BYTES = int(
    os.getenv("EXCEL_TO_CSV_MIN_SIZE_MB", "0")
) * 1024 * 1024


def _convert_excel_to_csv_if_possible(file_path, file_type):
    """
    Convert single-sheet Excel uploads to CSV for faster chunked ingestion.
    Returns the path to the file that should be processed.
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in {".xlsx", ".xlsm"}:
        return file_path

    if file_type in NON_CONVERTIBLE_EXCEL_TYPES:
        return file_path

    if os.path.getsize(file_path) < EXCEL_TO_CSV_MIN_SIZE_BYTES:
        return file_path

    csv_path = f"{os.path.splitext(file_path)[0]}.csv"
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    with open(csv_path, "w", newline="", encoding="utf-8") as out_file:
        writer = csv.writer(out_file)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if val is None else val for val in row])

    wb.close()
    return csv_path


@shared_task(bind=True)
def process_upload_file_task(
    self,
    file_path,
    file_type,
    user_id,
    data_owner_id,
    upload_log_id=None,
    date_str="",
    is_last=False,
    is_flipkart=False,
    batch_id="",
    batch_total=None,
):
    """
    Celery task that processes a single uploaded file.

    Parameters
    ----------
    file_path : str
        Absolute path to the uploaded file saved on disk.
    file_type : str
        One of: 'sales', 'spend', 'category', 'price',
        'fk_search_traffic', 'fk_category', 'fk_price',
        'fk_pla', 'fk_fba_stock', 'fk_flex_stock'.
    user_id : int
        ID of the logged-in user (for WebSocket notifications).
    data_owner_id : int
        ID of the data-owner user (main user) for DB associations.
    date_str : str
        Date string for sales files (DD-MM-YYYY).
    is_last : bool
        Whether this is the last file in the batch — triggers dashboard
        data generation and materialized view refresh.
    is_flipkart : bool
        Whether this file belongs to the Flipkart pipeline.
    """
    from apps.accounts.models import Users  # noqa: F401
    from django.conf import settings
    from apps.upload.models import UploadLog
    from apps.upload.amazon import (
        process_category_file,
        process_price_file,
        process_spend_file,
        process_sales_file,
        process_fba_stock_file,
        process_flex_stock_file,
    )
    from apps.upload.flipkart import (
        process_fk_fba_stock_file,
        process_fk_flex_stock_file,
        process_fk_inventory_file,
        process_fk_search_traffic,
        process_fk_category,
        process_fk_price,
        process_fk_pla,
    )
    _send_ws(user_id, f"Processing {file_type} file...", "processing")

    try:
        data_owner = Users.objects.get(pk=data_owner_id)
        files_to_cleanup = []
        keep_uploaded_files = bool(getattr(settings, "UPLOAD_KEEP_FILES", True))
        affected_dates = set()
        upload_log = None
        if upload_log_id:
            upload_log = UploadLog.objects.filter(pk=upload_log_id).first()
            if upload_log:
                upload_log.status = UploadLog.STATUS_PROCESSING
                upload_log.message = "Processing started."
                upload_log.save(update_fields=["status", "message", "updated_at"])

        processing_path = _convert_excel_to_csv_if_possible(file_path, file_type)
        if processing_path != file_path:
            # CSV converted from Excel is only an internal processing artifact.
            files_to_cleanup.append(processing_path)

        # Open the file from disk
        with open(processing_path, "rb") as fh:
            if file_type == "category":
                process_category_file(fh, data_owner)
            elif file_type == "price":
                process_price_file(fh, data_owner)
            elif file_type == "spend":
                affected_dates = process_spend_file(fh, data_owner) or set()
            elif file_type == "sales":
                affected_dates = process_sales_file(fh, date_str, data_owner) or set()
            elif file_type == "fba_stock":
                affected_dates = process_fba_stock_file(fh, data_owner) or set()
            elif file_type == "flex_stock":
                affected_dates = process_flex_stock_file(fh, data_owner) or set()
            elif file_type == "fk_search_traffic":
                affected_dates = process_fk_search_traffic(fh, data_owner) or set()
            elif file_type == "fk_category":
                process_fk_category(fh, data_owner)
            elif file_type == "fk_price":
                process_fk_price(fh, data_owner)
            elif file_type == "fk_pla":
                affected_dates = process_fk_pla(fh, data_owner) or set()
            elif file_type == "fk_fba_stock":
                affected_dates = process_fk_fba_stock_file(fh, data_owner) or set()
            elif file_type == "fk_flex_stock":
                affected_dates = process_fk_flex_stock_file(fh, data_owner) or set()
            elif file_type == "fk_inventory":
                affected_dates = process_fk_inventory_file(fh, data_owner) or set()
            else:
                raise ValueError(f"Unsupported file_type: {file_type}")

        # Clean up generated processing artifacts. Keep original uploads by default.
        if not keep_uploaded_files:
            files_to_cleanup.append(file_path)
        for path in files_to_cleanup:
            try:
                os.remove(path)
            except OSError:
                pass

        if batch_id:
            _send_ws(user_id, f"{file_type} processed successfully.", "partial")
            _mark_batch_task_complete(
                batch_id=batch_id,
                batch_total=batch_total,
                user_id=user_id,
                data_owner_id=data_owner.id,
                is_flipkart=is_flipkart,
                success=True,
                file_type=file_type,
                affected_dates=sorted(affected_dates),
            )
        elif is_last:
            _set_dashboard_refresh_status(
                data_owner.id,
                "processing",
                "Dashboard updating in process as per the new uploaded data.",
                timeout=3600,
            )
            refresh_dashboard_after_upload_task.delay(
                data_owner_id=data_owner.id,
                user_id=user_id,
                is_flipkart=is_flipkart,
                file_type=file_type,
                affected_dates=sorted(affected_dates),
                dashboard_refreshed=False,
            )
            _send_ws(
                user_id,
                "All files uploaded successfully. Dashboard update started.",
                "partial",
            )
        else:
            _send_ws(user_id, f"{file_type} processed successfully.", "partial")

        if upload_log:
            upload_log.status = UploadLog.STATUS_SUCCESS
            upload_log.message = "Processed successfully."
            upload_log.save(update_fields=["status", "message", "updated_at"])

        return {
            "status": "success",
            "file_type": file_type,
            "is_last": is_last,
        }
    except SoftTimeLimitExceeded:
        message = "Upload processing timed out. Please retry in smaller batches."
        logger.exception("[UploadTask] Soft time limit exceeded for %s", file_type)
        _send_ws(user_id, message, "error")

        if upload_log_id:
            try:
                upload_log = UploadLog.objects.filter(pk=upload_log_id).first()
                if upload_log:
                    upload_log.status = UploadLog.STATUS_ERROR
                    upload_log.message = message
                    upload_log.save(update_fields=["status", "message", "updated_at"])
            except Exception:
                logger.exception("[UploadTask] Failed updating timeout status.")

        if batch_id:
            _mark_batch_task_complete(
                batch_id=batch_id,
                batch_total=batch_total,
                user_id=user_id,
                data_owner_id=data_owner_id,
                is_flipkart=is_flipkart,
                success=False,
                file_type=file_type,
            )

        return {
            "status": "error",
            "file_type": file_type,
            "message": message,
        }


    except Exception as exc:
        logger.exception("[UploadTask] Error processing %s: %s", file_type, exc)
        _send_ws(user_id, f"Error processing file: {str(exc)}", "error")

        if upload_log_id:
            try:
                upload_log = UploadLog.objects.filter(pk=upload_log_id).first()
                if upload_log:
                    upload_log.status = UploadLog.STATUS_ERROR
                    upload_log.message = str(exc)
                    upload_log.save(update_fields=["status", "message", "updated_at"])
            except Exception:
                logger.exception("[UploadTask] Failed updating UploadLog status.")

        # Clean up generated CSV artifact and optionally original upload on failure.
        cleanup_candidates = [f"{os.path.splitext(file_path)[0]}.csv"]
        if not bool(getattr(settings, "UPLOAD_KEEP_FILES", True)):
            cleanup_candidates.append(file_path)
        for path in cleanup_candidates:
            try:
                os.remove(path)
            except OSError:
                pass

        if batch_id:
            _mark_batch_task_complete(
                batch_id=batch_id,
                batch_total=batch_total,
                user_id=user_id,
                data_owner_id=data_owner_id,
                is_flipkart=is_flipkart,
                success=False,
                file_type=file_type,
            )

        return {
            "status": "error",
            "file_type": file_type,
            "message": str(exc),
        }
